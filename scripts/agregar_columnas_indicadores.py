"""
Script para agregar las columnas "Fórmula de Cálculo" y "Explicación y Requisitos del Cálculo"
al archivo Excel de indicadores CESU 2025.

Uso:
    python scripts/agregar_columnas_indicadores.py

El script modifica el archivo:
    Manual indicadores-Lineamientos_y_Aspectos_CESU_2025.xlsx
"""

import os
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Ruta del archivo relativa al directorio raíz del repositorio
EXCEL_FILENAME = "Manual indicadores-Lineamientos_y_Aspectos_CESU_2025.xlsx"

# Nombres de las nuevas columnas
COL_FORMULA = "Fórmula de Cálculo"
COL_EXPLICACION = "Explicación y Requisitos del Cálculo"

# Fila de encabezados en todas las hojas
HEADER_ROW = 4

# Configuración por hoja: columna del indicador (1-based), columna del tipo, columna de descripción
SHEET_CONFIG = {
    "PROGRAMAS ACADÉMICOS": {
        "indicador_col": 11,   # K
        "tipo_col": 9,         # I
        "descripcion_col": 12, # L
        "aspecto_col": 8,      # H (descripción aspecto a evaluar)
        "first_data_row": 5,
    },
    "UNIDADES ACADÉMICAS": {
        "indicador_col": 9,    # I
        "tipo_col": 7,         # G
        "descripcion_col": 10, # J
        "aspecto_col": 6,      # F (descripción aspecto a evaluar)
        "first_data_row": 5,
    },
    "INSTITUCIONES DE ED. SUPERIOR ": {
        "indicador_col": 9,    # I
        "tipo_col": 7,         # G
        "descripcion_col": 10, # J
        "aspecto_col": 6,      # F (descripción aspecto a evaluar)
        "first_data_row": 5,
    },
}


# ---------------------------------------------------------------------------
# Lógica de generación de fórmulas y explicaciones
# ---------------------------------------------------------------------------

def inferir_formula_y_explicacion(indicador: str, tipo: str, descripcion: str, aspecto: str):
    """
    Infiere la fórmula de cálculo y la explicación a partir del texto del indicador
    y de la información de contexto disponible.

    Retorna una tupla (formula, explicacion).
    """
    ind = (indicador or "").strip()
    tip = (tipo or "").strip().lower()
    desc = (descripcion or "").strip()
    asp = (aspecto or "").strip()

    ind_lower = ind.lower()

    # -----------------------------------------------------------------------
    # 1. Detectar si el indicador ya contiene una fracción/ratio explícita
    # -----------------------------------------------------------------------
    if _contiene_fraccion(ind):
        formula = _formula_desde_fraccion(ind)
        explicacion = _explicacion_ratio(ind, desc, asp)
        return formula, explicacion

    # -----------------------------------------------------------------------
    # 2. Indicadores de porcentaje ("Porcentaje de ...")
    # -----------------------------------------------------------------------
    if ind_lower.startswith("porcentaje de"):
        formula = _formula_porcentaje(ind)
        explicacion = _explicacion_porcentaje(ind, desc, asp)
        return formula, explicacion

    # -----------------------------------------------------------------------
    # 3. Indicadores de cobertura
    # -----------------------------------------------------------------------
    if "cobertura" in ind_lower:
        formula = "(N° de elementos cubiertos / N° total de elementos objetivo) × 100"
        explicacion = _explicacion_cobertura(ind, desc, asp)
        return formula, explicacion

    # -----------------------------------------------------------------------
    # 4. Indicadores documentales / de logro no numérico
    # -----------------------------------------------------------------------
    doc_patterns = [
        "documento que demuestre",
        "un documento de",
        "documento de",
    ]
    if any(ind_lower.startswith(p) for p in doc_patterns):
        formula = "Verificación documental: 1 (cumple) o 0 (no cumple)"
        explicacion = _explicacion_documental(ind, desc, asp)
        return formula, explicacion

    # -----------------------------------------------------------------------
    # 5. Indicadores de conteo / N.º de / Nº de
    # -----------------------------------------------------------------------
    if re.match(r"^(n\.?º|nº|número|no\.)\s+de\s+", ind_lower):
        formula = _formula_conteo(ind, desc)
        explicacion = _explicacion_conteo(ind, desc, asp)
        return formula, explicacion

    # -----------------------------------------------------------------------
    # 6. Indicadores de porcentaje que empiezan con "%" en el texto
    # -----------------------------------------------------------------------
    if ind_lower.startswith("% de") or ind_lower.startswith("% "):
        formula = _formula_porcentaje(ind)
        explicacion = _explicacion_porcentaje(ind, desc, asp)
        return formula, explicacion

    # -----------------------------------------------------------------------
    # 7. Indicadores cuya descripción aclara el tipo de cálculo
    # -----------------------------------------------------------------------
    desc_lower = desc.lower()
    if "porcentaje" in desc_lower[:80]:
        formula = _formula_porcentaje_desde_desc(desc)
        explicacion = _explicacion_desde_descripcion(ind, desc, asp)
        return formula, explicacion

    # -----------------------------------------------------------------------
    # 8. Fallback: logro binario o índice compuesto
    # -----------------------------------------------------------------------
    if tip in ("logro", "impacto"):
        if "porcentaje" in ind_lower or "%" in ind:
            formula = "(Numerador del indicador / Total de referencia) × 100"
        else:
            formula = "Verificación cualitativa / Índice de logro: 0 a 100"
        explicacion = _explicacion_generica(ind, desc, asp, tip)
        return formula, explicacion

    formula = "(Valor parcial / Valor total de referencia) × 100"
    explicacion = _explicacion_generica(ind, desc, asp, tip)
    return formula, explicacion


# ---------------------------------------------------------------------------
# Funciones auxiliares para detección
# ---------------------------------------------------------------------------

def _contiene_fraccion(ind: str) -> bool:
    """Detecta si el indicador contiene explícitamente una fracción (A / B)."""
    # Busca patrones como: "Nº X / Nº total", "(No. A / No. B)", "N.º ... / N.º total"
    # Permite textos largos entre la referencia numérica y la barra
    return bool(re.search(r"(?:nº|n\.º|no\.|número)\s*.{0,200}/", ind, re.IGNORECASE | re.DOTALL))


def _formula_desde_fraccion(ind: str) -> str:
    """Extrae y construye la fórmula a partir de una fracción ya presente en el indicador."""
    # Limpiar el texto para construir la fórmula
    ind_clean = ind.replace("\n", " ").strip()
    # Extraer sólo la primera parte (antes de cualquier bloque adicional tipo "Permanencia:")
    primera_parte = re.split(r"\n\n|[A-ZÁÉÍÓÚÑ][a-záéíóúñ]+:", ind_clean)[0].strip()
    # Si tiene barra divisoria visible, usar esa expresión
    if "/" in primera_parte:
        partes = primera_parte.split("/", 1)
        num = partes[0].strip().rstrip("(").strip()
        den = partes[1].strip().rstrip(")").strip()
        # Si ya incluye factor × 100 evitar duplicarlo
        if "×" in den or "x 100" in den.lower():
            return f"({num}) / ({den})"
        return f"({num}) / ({den}) × 100"
    # Fallback
    if "×" in ind_clean or "x 100" in ind_clean.lower() or "x100" in ind_clean.lower():
        return f"({primera_parte[:150]}) × 100  [si no está ya multiplicado]"
    return f"({primera_parte[:150]}) × 100"


def _formula_porcentaje(ind: str) -> str:
    """Genera una fórmula de porcentaje extrapolando el numerador del texto."""
    # Intentar extraer la entidad del "porcentaje de X"
    match = re.match(r"(?:porcentaje|%)\s+de\s+(.+)", ind, re.IGNORECASE)
    if match:
        entidad = match.group(1).split("con")[0].split("que")[0].strip()
        entidad = entidad[:80].rstrip(" ,;")
        return f"(N° de {entidad} / N° total de referencia) × 100"
    return "(Numerador del indicador / Total de referencia) × 100"


def _formula_porcentaje_desde_desc(desc: str) -> str:
    """Construye fórmula de porcentaje extrayendo elementos de la descripción."""
    match = re.search(r"porcentaje de\s+(.+?)(?:que|con|sobre|/|\.)", desc[:200], re.IGNORECASE)
    if match:
        entidad = match.group(1).strip()
        return f"(N° de {entidad} / N° total de referencia) × 100"
    return "(Numerador del indicador / Total de referencia) × 100"


def _formula_conteo(ind: str, desc: str) -> str:
    """Genera fórmula para indicadores de conteo o tasa."""
    if "tasa" in ind.lower() or "tasa" in desc.lower():
        return "(N° de eventos en el período / Población base del período) × 100"
    return "Conteo directo: N° total de [elementos] en el período de referencia"


# ---------------------------------------------------------------------------
# Funciones auxiliares para explicaciones
# ---------------------------------------------------------------------------

def _explicacion_ratio(ind: str, desc: str, asp: str) -> str:
    partes = [
        "Fórmula de ratio/porcentaje:",
        f"  Numerador: Cantidad de elementos que cumplen la condición descrita en el indicador.",
        f"  Denominador: Cantidad total de elementos del universo de referencia.",
        "Datos necesarios:",
        "  • Registro o base de datos actualizada del período de reporte.",
        "  • Numerador y denominador obtenidos de la misma fuente y período.",
        "Pasos:",
        "  1. Definir el período de medición.",
        "  2. Contar el numerador (elementos que cumplen la condición).",
        "  3. Contar el denominador (total del universo).",
        "  4. Dividir y multiplicar por 100 para obtener el porcentaje.",
    ]
    if desc:
        partes.insert(1, f"Contexto: {desc[:200]}")
    return "\n".join(partes)


def _explicacion_porcentaje(ind: str, desc: str, asp: str) -> str:
    partes = [
        "Indicador de porcentaje:",
        "  Se calcula dividiendo la cantidad parcial (elementos que cumplen la condición) entre",
        "  el total del universo de referencia, y multiplicando por 100.",
        "Datos necesarios:",
        "  • Fuente primaria: registros académicos, sistemas de información institucional.",
        "  • Período de reporte: generalmente semestral o anual.",
        "  • Universo base: todos los elementos del conjunto (estudiantes, docentes, cursos, etc.).",
        "Pasos:",
        "  1. Establecer el período de medición (ej. semestre o año académico).",
        "  2. Identificar el universo total (denominador).",
        "  3. Filtrar los elementos que cumplen la condición del indicador (numerador).",
        "  4. Aplicar: (Numerador / Denominador) × 100.",
        "  5. Registrar y comparar con la línea base.",
    ]
    if desc:
        partes.insert(1, f"Contexto: {desc[:200]}")
    return "\n".join(partes)


def _explicacion_cobertura(ind: str, desc: str, asp: str) -> str:
    partes = [
        "Indicador de cobertura (porcentaje):",
        "  Mide qué fracción del universo objetivo ha sido alcanzada o cubierta.",
        "Datos necesarios:",
        "  • Listado completo del universo objetivo (denominador).",
        "  • Registro de los elementos efectivamente cubiertos (numerador).",
        "  • Fuente: bases de datos institucionales o matrices de seguimiento.",
        "Pasos:",
        "  1. Definir el universo objetivo total.",
        "  2. Contar los elementos efectivamente cubiertos o atendidos.",
        "  3. Aplicar: (N° cubiertos / N° total objetivo) × 100.",
        "  4. Verificar completitud de los registros antes del cálculo.",
    ]
    if desc:
        partes.insert(1, f"Contexto: {desc[:200]}")
    return "\n".join(partes)


def _explicacion_documental(ind: str, desc: str, asp: str) -> str:
    partes = [
        "Indicador de logro documental (binario):",
        "  Verifica la existencia y calidad de un documento o evidencia específica.",
        "  Valor: 1 = El documento existe y cumple los criterios; 0 = No existe o es insuficiente.",
        "Datos necesarios:",
        "  • El documento o evidencia referenciada en el indicador.",
        "  • Criterios de valoración: vigencia, coherencia, completitud y validación institucional.",
        "Pasos:",
        "  1. Identificar el documento requerido.",
        "  2. Verificar su existencia en el sistema de gestión documental.",
        "  3. Evaluar si cumple con los criterios de calidad establecidos.",
        "  4. Asignar valor 1 (cumple) o 0 (no cumple).",
        "  5. Documentar hallazgos y fecha de verificación.",
    ]
    if desc:
        partes.insert(1, f"Contexto: {desc[:200]}")
    return "\n".join(partes)


def _explicacion_conteo(ind: str, desc: str, asp: str) -> str:
    partes = [
        "Indicador de conteo / razón:",
        "  Cuantifica un número absoluto o calcula una tasa/razón entre dos grupos.",
        "Datos necesarios:",
        "  • Registros del sistema de información académica o administrativa.",
        "  • Período de referencia claramente definido.",
        "  • Definición operacional de la unidad de conteo.",
        "Pasos:",
        "  1. Definir el período de medición.",
        "  2. Establecer la definición exacta de los elementos a contar.",
        "  3. Realizar el conteo o el cociente entre grupos.",
        "  4. Registrar el resultado y comparar con períodos anteriores.",
    ]
    if desc:
        partes.insert(1, f"Contexto: {desc[:200]}")
    return "\n".join(partes)


def _explicacion_desde_descripcion(ind: str, desc: str, asp: str) -> str:
    partes = [
        "Indicador de porcentaje (según descripción):",
        f"  {desc[:300]}" if desc else "",
        "Datos necesarios:",
        "  • Fuente: sistema de información institucional y registros del período.",
        "  • Numerador: cantidad de elementos que cumplen la condición.",
        "  • Denominador: total del universo de referencia.",
        "Pasos:",
        "  1. Consultar el sistema de información para el período de reporte.",
        "  2. Filtrar el numerador según la condición del indicador.",
        "  3. Obtener el total del denominador.",
        "  4. Calcular: (Numerador / Denominador) × 100.",
    ]
    return "\n".join([p for p in partes if p])


def _explicacion_generica(ind: str, desc: str, asp: str, tipo: str) -> str:
    tipo_label = "logro" if "logro" in tipo else "impacto"
    partes = [
        f"Indicador de {tipo_label}:",
        f"  {desc[:300]}" if desc else f"  {ind[:200]}",
        "Datos necesarios:",
        "  • Fuentes institucionales: sistema de información académica, informes de gestión.",
        "  • Período de medición: según la frecuencia de actualización definida.",
        "Pasos:",
        "  1. Recopilar los datos del período de reporte.",
        "  2. Identificar el numerador (cantidad que cumple la condición) y el denominador (total).",
        "  3. Calcular el porcentaje o valor del indicador.",
        "  4. Registrar el resultado y comparar con la línea base y la meta.",
        "Consideraciones:",
        "  • Usar definiciones operacionales consistentes entre períodos.",
        "  • Documentar la fuente y fecha de captura de los datos.",
    ]
    return "\n".join([p for p in partes if p])


# ---------------------------------------------------------------------------
# Estilos de celda
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_ALIGN = Alignment(vertical="top", wrap_text=True)
DATA_FONT = Font(size=9)

THIN_BORDER_SIDE = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(
    left=THIN_BORDER_SIDE,
    right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE,
    bottom=THIN_BORDER_SIDE,
)

COL_FORMULA_WIDTH = 45
COL_EXPLICACION_WIDTH = 70


# ---------------------------------------------------------------------------
# Función principal de procesamiento
# ---------------------------------------------------------------------------

def procesar_hoja(ws, config: dict):
    """
    Agrega las dos nuevas columnas a una hoja de trabajo (idempotente: si ya existen las sobreescribe).

    Args:
        ws: objeto worksheet de openpyxl
        config: diccionario con la configuración de columnas de la hoja
    """
    # Buscar si las columnas ya existen (idempotencia)
    col_formula_idx = None
    col_explicacion_idx = None
    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(HEADER_ROW, col).value
        if header_val == COL_FORMULA:
            col_formula_idx = col
        elif header_val == COL_EXPLICACION:
            col_explicacion_idx = col

    if col_formula_idx and col_explicacion_idx:
        print(f"  Hoja '{ws.title}': columnas ya existen, sobreescribiendo.")
    else:
        # Determinar las columnas de destino (después de la última columna con datos reales,
        # ignorando las columnas que ya podríamos haber agregado previamente)
        last_col = _encontrar_ultima_columna_original(ws)
        col_formula_idx = last_col + 1
        col_explicacion_idx = last_col + 2
        print(f"  Hoja '{ws.title}': última columna original = {get_column_letter(last_col)} ({last_col})")

    print(f"  Escribiendo: {get_column_letter(col_formula_idx)} = '{COL_FORMULA}'")
    print(f"               {get_column_letter(col_explicacion_idx)} = '{COL_EXPLICACION}'")

    # Escribir encabezados
    _escribir_encabezado(ws, HEADER_ROW, col_formula_idx, COL_FORMULA)
    _escribir_encabezado(ws, HEADER_ROW, col_explicacion_idx, COL_EXPLICACION)

    # Ajustar ancho de columnas
    ws.column_dimensions[get_column_letter(col_formula_idx)].width = COL_FORMULA_WIDTH
    ws.column_dimensions[get_column_letter(col_explicacion_idx)].width = COL_EXPLICACION_WIDTH

    # Procesar filas de datos
    indicador_col = config["indicador_col"]
    tipo_col = config["tipo_col"]
    descripcion_col = config["descripcion_col"]
    aspecto_col = config["aspecto_col"]
    first_data_row = config["first_data_row"]

    filas_procesadas = 0
    for row_idx in range(first_data_row, ws.max_row + 1):
        indicador_val = ws.cell(row_idx, indicador_col).value
        tipo_val = ws.cell(row_idx, tipo_col).value
        descripcion_val = ws.cell(row_idx, descripcion_col).value
        aspecto_val = ws.cell(row_idx, aspecto_col).value

        if not indicador_val and not tipo_val:
            # Fila sin indicador: dejar celdas vacías con borde
            _escribir_dato(ws, row_idx, col_formula_idx, "")
            _escribir_dato(ws, row_idx, col_explicacion_idx, "")
            continue

        formula, explicacion = inferir_formula_y_explicacion(
            str(indicador_val) if indicador_val else "",
            str(tipo_val) if tipo_val else "",
            str(descripcion_val) if descripcion_val else "",
            str(aspecto_val) if aspecto_val else "",
        )

        _escribir_dato(ws, row_idx, col_formula_idx, formula)
        _escribir_dato(ws, row_idx, col_explicacion_idx, explicacion)
        filas_procesadas += 1

    print(f"  → {filas_procesadas} indicadores procesados.")


def _encontrar_ultima_columna_original(ws) -> int:
    """
    Retorna el índice de la última columna con datos originales, ignorando
    cualquier columna cuyo encabezado sea COL_FORMULA o COL_EXPLICACION.
    """
    # Columnas que debemos ignorar (nuestras propias columnas si ya existen)
    ignorar = set()
    for col in range(1, ws.max_column + 1):
        h = ws.cell(HEADER_ROW, col).value
        if h in (COL_FORMULA, COL_EXPLICACION):
            ignorar.add(col)

    last_col = 1
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and cell.column not in ignorar:
                if cell.column > last_col:
                    last_col = cell.column
    return last_col


def _escribir_encabezado(ws, row: int, col: int, texto: str):
    cell = ws.cell(row, col, texto)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = THIN_BORDER


def _escribir_dato(ws, row: int, col: int, texto: str):
    cell = ws.cell(row, col, texto)
    cell.font = DATA_FONT
    cell.alignment = DATA_ALIGN
    cell.border = THIN_BORDER


# ---------------------------------------------------------------------------
# Entrada principal
# ---------------------------------------------------------------------------

def main():
    # Buscar el archivo Excel en el directorio de trabajo o en el padre
    script_dir = os.path.dirname(os.path.abspath(__file__))
    repo_root = os.path.dirname(script_dir)

    excel_path = None
    for candidate in [
        os.path.join(repo_root, EXCEL_FILENAME),
        os.path.join(os.getcwd(), EXCEL_FILENAME),
        EXCEL_FILENAME,
    ]:
        if os.path.exists(candidate):
            excel_path = candidate
            break

    if not excel_path:
        raise FileNotFoundError(
            f"No se encontró el archivo '{EXCEL_FILENAME}'. "
            "Ejecuta este script desde la raíz del repositorio."
        )

    print(f"Leyendo: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)

    for sheet_name, config in SHEET_CONFIG.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\nProcesando hoja: '{sheet_name}'")
            procesar_hoja(ws, config)
        else:
            print(f"Advertencia: hoja '{sheet_name}' no encontrada, se omite.")

    print(f"\nGuardando: {excel_path}")
    wb.save(excel_path)
    print("¡Listo! El archivo fue guardado correctamente.")


if __name__ == "__main__":
    main()
